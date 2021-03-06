import openpyxl
import datetime
import Data

def GetWorkbook():
	file_path = input("Wprowadz lokalizacje pliku\n")
	if file_path == "":
		file_path='Zadanie praca.xlsx'
		print('Wybrano domyslna sciezke: ',end=file_path+'\n')
	try:
		wb = openpyxl.load_workbook(filename = file_path,data_only=True)
	except IOError as ex:
		print("Niepoprawna sciezka")
	return wb

def PickSheet(wb):
	options = [str(x+1)+'. '+wb.sheetnames[x] for x in range(len(wb.sheetnames))]
	print("Wybierz arkusz")
	print(options)
	picked_number = input("Wybierz numer\n")
	try:
		data_sheet_index = int(picked_number)
		data_sheet_index = data_sheet_index - 1
	except ValueError:
		print('Przyjeto domyslnie pierwszy arkusz')
		data_sheet_index=0
	sheet = wb[wb.sheetnames[data_sheet_index]]
	return sheet


def GetHeaders(sheet,maxrow=1000,maxcol=1000):
    headers=[]
    flag_found_headers=False
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=maxrow+1, max_col=maxcol+1):
        for cell in row:
            if cell.value is not None:
                flag_found_headers=True
                headers.append(cell.value)
        if flag_found_headers == True:
            return headers
    if flag_found_headers == False:
        raise Exception('Brak poprawnych danych w arkuszu')

def CopyTable(sheet,maxrow=1000,maxcol=1000):
    """
    Returns table from the worksheet as a list of lists. The first sublist are headers. The rest of sublists is data.
    """
    headers=GetHeaders(sheet,maxrow,maxcol)
    data_length=len(headers)
    table_copy=[]
    flag_found_stm=False
    for row in sheet.iter_rows(min_row=1, min_col=1, max_row=maxrow+1, max_col=maxcol+1):
        in_data_pos=0
        for cell in row:
            if cell.value is not None or 0 < in_data_pos < data_length :
                flag_found_stm=True
                in_data_pos+=1
                table_copy.append(cell.value)
        if flag_found_stm == True and in_data_pos == 0:
            return RetriveHeadersAndData(table_copy,data_length)
    print('Nie wszystkie dane zostaly sczytane: zwiekszam zakres wyszukiwan. To moze troche potrwac...')
    if maxrow == 1000000 and maxcol == 1000000:
        raise Exception("W podany plik jest pusty")
    return CopyTable(sheet,maxrow*10,maxcol*10)
 
def RetriveHeadersAndData(table,headers_length):
    headers=table[:headers_length]
    for x in range(len(headers)):
         headers[x]=headers[x].rstrip()   #usuwam wszytkie spacje z prawej strony
    data=[]
    i=headers_length
    while i+1< len(table):
        data.append(table[i:i+headers_length])
        i+=headers_length     
    return headers,data

if __name__=='__main__':
    wb=GetWorkbook()
    sheet=PickSheet(wb)
    headers , data=CopyTable(sheet)
    dictionary=dict()
    all_data=[]
    for x in range(len(data)):
        dictionary = dict(zip(headers, data[x]))
        all_data.append(Data.Data(dictionary))
        print(all_data[x])
        print(all_data[x].AvgCost())
 
